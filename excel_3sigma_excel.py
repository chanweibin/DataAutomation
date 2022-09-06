import argparse, sys, os
import glob
import shutil
import pandas as pd
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.chart import (
    LineChart,
    Reference,
)
import data_df_repos as dfr
import io_save_file as iosf
import io_read_file as iorf
from datetime import datetime

# data file (txt) to dataframe
def file_to_dataframe(inputfile_loc):
    df = pd.DataFrame({'Test Name':[],'Lo':[],'Result':[],'Hi':[],'%':[],'P/F':[]})

    file1 = open(inputfile_loc, 'r')
    Lines = file1.readlines()

    is_data = False
    for line in Lines:
        # indicate start of testpoints (reset df to get latest test)
        if "OUTPUT#" in line and "TESTS" in line:
            df = pd.DataFrame(columns=df.columns)
            is_data = True
            continue
        # indicate end of testpoints
        if "*** TEST RUN" in line:
            is_data = False
            continue
        # save line to dataframe if line is testpoint
        if is_data == True:
            # ignore lines with "Waveform for ..."
            if "Waveform for" in line or line == "":
                continue
            # split line by space
            data = line.split()
            # recombine testname if case it has space
            testname = data[0]
            for index in range(1,data.index(data[-6])+1):
                testname += " " + data[index]
            # Concatenate line into df
            new_row = pd.DataFrame({'Test Name':[testname],'Lo':[data[-5]],'Result':[data[-4]],'Hi':[data[-3]],'%':[data[-2]],'P/F':[data[-1]]})
            df = pd.concat([df,new_row],ignore_index=True)
    if df.empty:
        print(f"No data found in file: {inputfile_loc} ...\n")
    return df

# Plot graph in excel
def plot_sigma_graph(worksheet,cols_list,start,length,title,rowpos,colpos,width=20,height=10):
    # Create a chart object.
    chart = LineChart()
    chart.width = width
    chart.height = height
    chart.title = title
    # Data
    start += 1
    # Go thru each column for x-axis
    for col in cols_list:
        data = Reference(worksheet, min_col=col, min_row=start, max_col=col, max_row=start + length)
        chart.add_data(data, titles_from_data=True)
    # y-axis
    name = Reference(worksheet, min_col=1, min_row=start + 1, max_col=1, max_row=start + length)
    chart.set_categories(name)
    # Graph properties
    chart.x_axis.tickLblPos = "low"
    s1 = chart.series[-1]
    s1.graphicalProperties.line.solidFill = "FF0000"
    s2 = chart.series[-2]
    s2.graphicalProperties.line.solidFill = "FF0000"
    # Insert the chart into the worksheet
    letter = get_column_letter(colpos)
    pos = letter + str(rowpos)
    rowpos = rowpos + 20 # leave 20 cells as start for next graph
    worksheet.add_chart(chart,pos)
    return rowpos

def create_subtable_keyword(DF,keyword_list,writer,title,startrowindex=1,lastpos=1):
    result = dfr.keyword_filter_row(DF,'Test Name',keyword_list)
    if len(result.index) == 0:
        return lastpos, startrowindex
    result = result.rename(columns={'Test Name':title})
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=startrowindex)
    worksheet = writer.book['Data Analysis']
    lastpos = plot_sigma_graph(worksheet,range(2,len(df.columns)+1),startrowindex,result.shape[0],title,lastpos,result.shape[1] + 2)
    startrowindex += result.shape[0] + 2
    return lastpos, startrowindex

def create_subtable_keywords(DF,keyword_list,writer,title,startrowindex=1,lastpos=1):
    result = dfr.keywords_filter_row(DF,'Test Name',keyword_list)
    if len(result.index) == 0:
        return lastpos, startrowindex
    result = result.rename(columns={'Test Name':title})
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=startrowindex)
    worksheet = writer.book['Data Analysis']
    lastpos = plot_sigma_graph(worksheet,range(2,len(df.columns)+1),startrowindex,result.shape[0],title,lastpos,result.shape[1] + 2)
    startrowindex += result.shape[0] + 2
    return lastpos, startrowindex

def txt_file():
    # Read the first file
    current_file_loc = all_files[0]
    df_temp = file_to_dataframe(current_file_loc)
    df = pd.DataFrame()
    df['Test Name'] = df_temp['Test Name']
    name = os.path.basename(all_files[0])
    folder_and_name = all_files[0].replace(input_file_loc + "\\",'')
    df[folder_and_name] = df_temp['%'].astype(float)

    # Move first file to "Done" folder
    destination = all_files[0].replace(name,'') + "Done"
    try:
        shutil.move(all_files[0] , destination + "\\" + name)
    except FileNotFoundError:
        os.mkdir(destination)
        shutil.move(all_files[0] , destination + "\\" + name)
    except:
        pass

    # Read every other files
    for file in range(1,len(all_files)):
        current_file_loc = all_files[file]
        df_temp = file_to_dataframe(current_file_loc)
        if df_temp.empty: continue
        # Merge the read with the main DF according to matching "Test Name"
        df = df.merge(df_temp[['Test Name','%']],on='Test Name').drop_duplicates()
        name = os.path.basename(all_files[file])
        folder_and_name = all_files[file].replace(input_file_loc + "\\",'')
        df['%'] = df['%'].astype(float)
        df = df.rename(columns={'%':folder_and_name}) # Rename column header to filename
        
        # Move file to "Done" folder
        destination = all_files[file].replace(name,'') + "Done"
        try:
            shutil.move(all_files[file] , destination + "\\" + name)
        except FileNotFoundError:
            os.mkdir(destination)
            shutil.move(all_files[file] , destination + "\\" + name)
        except:
            pass

    # Add the '+3 Sigma' & '-3 Sigma' from the latest file in the '3Sigma' folder
    df = df.merge(df_sigma[['Test Name','+3 Sigma','-3 Sigma']],on='Test Name').drop_duplicates()

    return df

def excel_file():
    # Read the first file
    current_file_loc = all_files[0]
    xls = pd.ExcelFile(current_file_loc, engine="openpyxl")
    df_temp = pd.read_excel(xls,0,header=13,usecols="B,D")
    df = pd.DataFrame()
    df['Limit Name'] = df_temp['Limit Name']
    name = os.path.basename(all_files[0])
    folder_and_name = all_files[0].replace(input_file_loc + "\\",'')
    df[folder_and_name] = df_temp['Actual Value'].astype(float)

    # Move first file to "Done" folder
    destination = all_files[0].replace(name,'') + "Done"
    try:
        shutil.move(all_files[0] , destination + "\\" + name)
    except FileNotFoundError:
        os.mkdir(destination)
        shutil.move(all_files[0] , destination + "\\" + name)
    except:
        pass

    # Read every other files
    for file in range(1,len(all_files)):
        current_file_loc = all_files[file]
        xls = pd.ExcelFile(current_file_loc, engine="openpyxl")
        df_temp = pd.read_excel(xls,0,header=13,usecols="B,D")
        if df_temp.empty: continue
        # Merge the read with the main DF according to matching "Test Name"
        df = df.merge(df_temp[['Limit Name','Actual Value']],on='Limit Name')
        df.drop_duplicates(subset=['Limit Name'], keep='first', inplace=True, ignore_index=True)
        name = os.path.basename(all_files[file])
        folder_and_name = all_files[file].replace(input_file_loc + "\\",'')
        df['Actual Value'] = df['Actual Value'].astype(float)
        df = df.rename(columns={'Actual Value':folder_and_name}) # Rename column header to filename
        
        # Move file to "Done" folder
        destination = all_files[file].replace(name,'') + "Done"
        try:
            shutil.move(all_files[file] , destination + "\\" + name)
        except FileNotFoundError:
            os.mkdir(destination)
            shutil.move(all_files[file] , destination + "\\" + name)
        except:
            pass

    # Add the '+3 Sigma' & '-3 Sigma' from the latest file in the '3Sigma' folder
    df = df.merge(df_sigma[['Limit Name','+3 Sigma','-3 Sigma']],on='Limit Name').drop_duplicates()
    df = df.rename(columns={'Limit Name':'Test Name'})
    return df

cwd = os.getcwd()
scriptName = sys.argv[0]
parser = argparse.ArgumentParser(description=str("Inputing arguments for " + scriptName))
parser.add_argument('-input', metavar="Input file location", help="Source file location", default=cwd)
args = parser.parse_args()

if args.input:
    input_file_loc = args.input

#input_file_loc = "C:\\Local_Storage\\3_Sigma"

# Get 3 Sigma from latest file
list_of_files = glob.glob(input_file_loc + "\\3Sigma\\*")
latest_file = max(list_of_files,key=os.path.getctime)
df_sigma = iorf.excel_to_dataframe(latest_file)

# Read every file in every folder other than the 3 stated
all_files = []
for dir in os.listdir(input_file_loc):
    if os.path.isdir(input_file_loc + "\\" + dir) and not (dir == "Calc_3Sigma" or dir == "3Sigma" or dir == "Result"):
        path = input_file_loc + "\\" + dir
        for item in os.listdir(path):
            if os.path.isfile(path + "\\" + item):
                all_files.append(path + "\\" + item)

# End program if no files detected
if not all_files:
    print(f"File not found in the folder: {input_file_loc} ...\nExitting!")
    sys.exit()

if (all_files[0].endswith(".xlsx")):
    df = excel_file()
else:
    df = txt_file()

# Output to excel
output_file_name = "Compare_" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"
output_full_name = input_file_loc + '\\' + 'Result' + '\\' + output_file_name

writer = pd.ExcelWriter(output_full_name, engine='openpyxl')
book = writer.book

whole = df

cnt = 1
lastpos = 2
keyword_list = ['Acal','CheckModel','CheckSerialNumber','CheckFWRev','SelfTest','Humidity Test','Temperature Test',\
    'V Prog Accuracy','V Rdbk Accuracy','I Prog Accuracy','I Rdbk Accuracy','VoltageLineRegulation','VoltageLoadRegulation','CV_prg_up','CV_prg_dn',\
    'CurrentLineRegulation','CurrentLoadRegulation','TransientResponseLoading','TransientResponseUnloading','UP','DOWN','Noise-Range','OVP Accuracy']

lastpos, cnt = create_subtable_keywords(df,["Acal","CheckModel","CheckSerialNumber","CheckFWRev","SelfTest"],writer,'Acal,CheckModel,CheckSerialNumber,CheckFWRev,SelfTest',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keywords(df,["Humidity Test","Temperature Test"],writer,'Humidity Test","Temperature Test',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keyword(df,"V Prog Accuracy",writer,'V Prog Accuracy',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keyword(df,"V Rdbk Accuracy",writer,'V Rdbk Accuracy',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keyword(df,"I Prog Accuracy",writer,'I Prog Accuracy',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keyword(df,"I Rdbk Accuracy",writer,'I Rdbk Accuracy',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keywords(df,["VoltageLineRegulation","VoltageLoadRegulation"],writer,'Voltage Load\Line Regulation',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keywords(df,["CurrentLineRegulation","CurrentLoadRegulation"],writer,'Current Load\Line Regulation',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keywords(df,["TransientResponseLoading","TransientResponseUnloading"],writer,'Transient Response',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keyword(df,"UP",writer,'UP',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keyword(df,"DOWN",writer,'DOWN',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keyword(df,"Noise-Range",writer,'Noise Test',startrowindex=cnt,lastpos=lastpos)
lastpos, cnt = create_subtable_keyword(df,"OVP Accuracy",writer,'OVP Accuracy',startrowindex=cnt,lastpos=lastpos)

copy = whole
whole = whole.rename(columns={'Test Name':'Name'})
whole = whole[~whole.Name.str.contains('|'.join(keyword_list))]
if len(whole.index) != 0:
    whole = whole.rename(columns={'Name':'Others'})
    whole.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=cnt)
    worksheet = writer.book['Data Analysis']
    lastpos = plot_sigma_graph(worksheet,range(2,len(df.columns)+1),cnt,whole.shape[0],'Others',lastpos,whole.shape[1] + 2)
    cnt += whole.shape[0] + 2

writer.save()
print("saving file ...")

