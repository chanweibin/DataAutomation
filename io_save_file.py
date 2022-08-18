from fileinput import filename
import pandas as pd 
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import os, re, sys


# * ================================================ Sanity Check ================================================


# * =============================================== Save Dataframe =============================================

def dataframe_to_excel(df, output_full_path, sheetName="Sheet1"):
    try:
        output_file_name = output_full_path.split("\\")[-1]
        print(f" ================ Output File : {output_file_name} ================ ")
        print(f"File location : {output_full_path}\n")

        print(f"Checking if {output_file_name} exists ...")
        if not os.path.exists(output_full_path):
            print("File does not exist, creating new excel file now ...")
            writer = pd.ExcelWriter(output_full_path, engine='xlsxwriter')
            df.to_excel(writer, sheet_name=sheetName, index=False, header=True)
        # append excel sheet if excel exists
        else:
            print("File exists, appending data to exising file ...")
            excelWorkbook = load_workbook(output_full_path)
            writer = pd.ExcelWriter(output_full_path, engine='openpyxl')
            writer.book = excelWorkbook
            df.to_excel(writer, sheet_name=sheetName, index=False, header=True)
        writer.save()
        return print("... done\n")
        

    except Exception as e:
        raise Exception("Error at WriteExcel: " + str(e))



def append_to_sheet(df, output_full_path, sheetName,start_col=0,start_row=0):
    excelWorkbook = load_workbook(output_full_path)
    writer = pd.ExcelWriter(output_full_path, engine='openpyxl')
    writer.book = excelWorkbook
    df.to_excel(writer, sheet_name=sheetName, index=False, header=True, startcol=start_col, startrow=start_row)
    writer.save()
    print("saving file ...")



def create_excel_file(full_path):
    wb = Workbook(full_path)
    filename = full_path.split("\\")[-1]
    wb.save(filename)
    return