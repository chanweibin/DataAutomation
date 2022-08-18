import argparse, sys, os
import glob
import shutil
import pandas as pd
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.chart import (
    LineChart,
    Reference,
)
import data_df_repos as dfr
import io_save_file as iosf
import io_read_file as iorf
from datetime import datetime

def file_to_dataframe(inputfile_loc):
    # df = pd.read_fwf("C:\\Users\\weipong8\\Downloads\\OLYFTS8 12A\\C12002770.8")
    df = pd.DataFrame({'Test Name':[],'Lo':[],'Result':[],'Hi':[],'%':[],'P/F':[]})

    file1 = open(inputfile_loc, 'r')
    Lines = file1.readlines()

    is_data = False
    for line in Lines:
        if "OUTPUT#" in line and "TESTS" in line:
            df = pd.DataFrame(columns=df.columns)
            is_data = True
            continue
        if "*** TEST RUN" in line:
            is_data = False
            continue
        if is_data == True:
            test = line
            if "Waveform for" in line or line == "":
                continue
            data = line.split()
            testname = data[0]
            for index in range(1,data.index(data[-6])+1):
                testname += " " + data[index]
            new_row = pd.DataFrame({'Test Name':[testname],'Lo':[data[-5]],'Result':[data[-4]],'Hi':[data[-3]],'%':[data[-2]],'P/F':[data[-1]]})
            df = pd.concat([df,new_row],ignore_index=True)
    if df.empty:
        print(f"No data found in file: {inputfile_loc} ...\n")
    return df

def plot_sigma_graph(worksheet,cols_list,start,length,title,rowpos,colpos,width=40,height=20):
    # Create a chart object.
    chart = LineChart()
    chart.width = width
    chart.height = height
    chart.title = title
    # Data
    start += 1
    for col in cols_list:
        data = Reference(worksheet, min_col=col, min_row=start, max_col=col, max_row=start + length)
        chart.add_data(data, titles_from_data=True)
    name = Reference(worksheet, min_col=1, min_row=start + 1, max_col=1, max_row=start + length)
    chart.set_categories(name)
    chart.x_axis.tickLblPos = "low"
    s1 = chart.series[-1]
    s1.graphicalProperties.line.solidFill = "FF0000"
    s2 = chart.series[-2]
    s2.graphicalProperties.line.solidFill = "FF0000"
    # Insert the chart into the worksheet
    letter = get_column_letter(colpos)
    pos = letter + str(rowpos)
    rowpos = rowpos + 40
    worksheet.add_chart(chart,pos)
    return rowpos

cwd = os.getcwd()
scriptName = sys.argv[0]
parser = argparse.ArgumentParser(description=str("Inputing arguments for " + scriptName))
parser.add_argument('-input', metavar="Input file location", help="Source file location", default=cwd)
args = parser.parse_args()

if args.input:
    input_file_loc = args.input

list_of_files = glob.glob(input_file_loc + "\\3Sigma\\*")
latest_file = max(list_of_files,key=os.path.getctime)

df_sigma = iorf.excel_to_dataframe(latest_file)

all_files = []
for dir in os.listdir(input_file_loc):
    if os.path.isdir(input_file_loc + "\\" + dir) and not (dir == "Calc_3Sigma" or dir == "3Sigma" or dir == "Result"):
        path = input_file_loc + "\\" + dir
        for item in os.listdir(path):
            if os.path.isfile(path + "\\" + item) and not item.endswith(".xlsx"):
                all_files.append(path + "\\" + item)

if not all_files:
    print(f"File not found in the folder: {input_file_loc} ...\nExitting!")
    sys.exit()

current_file_loc = all_files[0]
df_temp = file_to_dataframe(current_file_loc)
df = pd.DataFrame()
df['Test Name'] = df_temp['Test Name']
name = os.path.basename(all_files[0])
folder_and_name = all_files[0].replace(input_file_loc + "\\",'')
df[folder_and_name] = df_temp['%'].astype(float)

destination = all_files[0].replace(name,'') + "Done"
try:
    shutil.move(all_files[0] , destination + "\\" + name)
except FileNotFoundError:
    os.mkdir(destination)
    shutil.move(all_files[0] , destination + "\\" + name)
except:
    pass

for file in range(1,len(all_files)):
    current_file_loc = all_files[file]
    df_temp = file_to_dataframe(current_file_loc)
    if df_temp.empty: continue
    df = df.merge(df_temp[['Test Name','%']],on='Test Name').drop_duplicates()
    name = os.path.basename(all_files[file])
    folder_and_name = all_files[file].replace(input_file_loc + "\\",'')
    df['%'] = df['%'].astype(float)
    df = df.rename(columns={'%':folder_and_name})
    destination = all_files[file].replace(name,'') + "Done"
    try:
        shutil.move(all_files[file] , destination + "\\" + name)
    except FileNotFoundError:
        os.mkdir(destination)
        shutil.move(all_files[file] , destination + "\\" + name)
    except:
        pass

df = df.merge(df_sigma[['Test Name','+3 Sigma','-3 Sigma']],on='Test Name').drop_duplicates()

output_file_name = "Compare_" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"
output_full_name = input_file_loc + '\\' + 'Result' + '\\' + output_file_name
iosf.dataframe_to_excel(df,output_full_name,"Data Analysis")
# pd.set_option("display.max_rows",None,"display.max_columns",None)
# print(df)

excelWorkbook = load_workbook(output_full_name)
writer = pd.ExcelWriter(output_full_name, engine='openpyxl')
writer.book = excelWorkbook
worksheet = excelWorkbook['Data Analysis']

plot_sigma_graph(worksheet,range(2,len(df.columns)+1),0,df.shape[0],"Compare",2,df.shape[1] + 2)

writer.save()
print("saving file ...")
