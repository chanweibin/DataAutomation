import argparse, sys, os
from concurrent.futures.process import _chain_from_iterable_of_lists
from ntpath import join
import shutil
from unittest import result
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.chart import (
    LineChart,
    Reference,
)
import io_save_file as iosf
import data_df_repos as dfr
from datetime import datetime

def mover(source, destination):
    files = [x for x in os.listdir(source) if x.endswith((".csv",".xlsx"))]
    for file in range(len(files)):
        filename = files[file]
        try:
            shutil.move(source + "\\" + filename , destination + "\\" + filename)
        except FileNotFoundError:
            os.mkdir(destination)
            shutil.move(source + "\\" + filename , destination + "\\" + filename)
        except:
            pass

def plot_sigma_graph(worksheet,cols_list,start,length,title,rowpos,colpos,width=45,height=20):
    # Create a chart object.
    chart = LineChart()
    chart.width = width
    chart.height = height
    chart.title = title
    # Data
    start += 1
    for col in cols_list:
        data = Reference(worksheet, min_col=col, min_row=start, max_col=col, max_row=start + length + 1)
        chart.add_data(data, titles_from_data=True)
    name = Reference(worksheet, min_col=1, min_row=start, max_col=1, max_row=start + length + 1)
    chart.set_categories(name)
    chart.x_axis.tickLblPos = "low"
    # Insert the chart into the worksheet
    letter = get_column_letter(colpos)
    pos = letter + str(rowpos)
    rowpos = rowpos + 40
    worksheet.add_chart(chart,pos)
    return rowpos

def create_subtable_keyword(DF1,DF2,keyword_list,writer,title,Parent3,Parent3_word="",startrowindex=1,lastpos=1):
    DF1_temp = dfr.keyword_filter_row(DF1,'Name',keyword_list)
    DF2_temp = dfr.keyword_filter_row(DF2,'Name',keyword_list)
    if Parent3:
        DF1_temp = dfr.keyword_filter_row(DF1_temp,'Parent3',Parent3_word)
        DF2_temp = dfr.keyword_filter_row(DF2_temp,'Parent3',Parent3_word)
    result = DF1_temp.merge(DF2_temp,on='Name',suffixes=("_(Eval 1)","_(Eval 2)"))
    if len(result.index) == 0:
        return lastpos, startrowindex
    result = result.rename(columns={'Name':title})
    result.insert(DF1_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=startrowindex)
    col_names = ['Mean_(Eval 1)','Mean_(Eval 2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,startrowindex,result.shape[0],title,lastpos,result.shape[1] + 1)
    startrowindex += result.shape[0] + 2
    return lastpos, startrowindex

def create_subtable_keywords(DF1,DF2,keyword_list,writer,title,Parent3,Parent3_word="",startrowindex=1,lastpos=1):
    DF1_temp = dfr.keywords_filter_row(DF1,'Name',keyword_list)
    DF2_temp = dfr.keywords_filter_row(DF2,'Name',keyword_list)
    if Parent3:
        DF1_temp = dfr.keyword_filter_row(DF1_temp,'Parent3',Parent3_word)
        DF2_temp = dfr.keyword_filter_row(DF2_temp,'Parent3',Parent3_word)
    result = DF1_temp.merge(DF2_temp,on='Name',suffixes=("_(Eval 1)","_(Eval 2)"))
    if len(result.index) == 0:
        return lastpos, startrowindex
    result = result.rename(columns={'Name':title})
    result.insert(DF1_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=startrowindex)
    col_names = ['Mean_(Eval 1)','Mean_(Eval 2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,startrowindex,result.shape[0],title,lastpos,result.shape[1] + 1)
    startrowindex += result.shape[0] + 2
    return lastpos, startrowindex

cwd = os.getcwd()
scriptName = sys.argv[0]
parser = argparse.ArgumentParser(description=str("Inputing arguments for " + scriptName))
parser.add_argument('-input', metavar="Input file location", help="Source file location", default="C:\Local_Storage\Data")
args = parser.parse_args()

if args.input:
    input_file_loc = args.input + "\\"

dir_path = os.path.dirname(os.path.realpath(__file__))
txt = 'py '+ dir_path +'\init.py -input '

mover(input_file_loc + 'Compiled\\Eval_1', input_file_loc + 'Compiled\\Eval_1\\Transferred')
temp = txt + input_file_loc + 'Eval_1' + ' -output ' + input_file_loc + 'Compiled\\' + 'Eval_1\\'
os.system(temp)
mover(input_file_loc + 'Eval_1', input_file_loc + 'Eval_1\\Transferred')

mover(input_file_loc + 'Compiled\\Eval_2', input_file_loc + 'Compiled\\Eval_2\\Transferred')
temp = txt + input_file_loc + 'Eval_2' + ' -output ' + input_file_loc + 'Compiled\\' + 'Eval_2\\'
os.system(temp)
mover(input_file_loc + 'Eval_2', input_file_loc + 'Eval_2\\Transferred')

output_file_name = "Balsa_" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"
output_full_name = input_file_loc + 'Result' + '\\' + output_file_name

DF2 = pd.DataFrame()
DF3 = pd.DataFrame()
folderpath = input_file_loc + 'Compiled\\' + 'Eval_1'
files = [x for x in os.listdir(folderpath) if x.endswith((".csv",".xlsx"))]
if len(files) > 0:
    DF2 = pd.read_excel(folderpath + '\\' + files[0])
    iosf.dataframe_to_excel(DF2,output_full_name,"Eval_1")

folderpath = input_file_loc + 'Compiled\\' + 'Eval_2'
files = [x for x in os.listdir(folderpath) if x.endswith((".csv",".xlsx"))]
if len(files) > 0:
    DF3 = pd.read_excel(folderpath + '\\' + files[0])
    iosf.dataframe_to_excel(DF3,output_full_name,"Eval_2")

if not DF2.empty and not DF3.empty :
    excelWorkbook = load_workbook(output_full_name)
    writer = pd.ExcelWriter(output_full_name, engine='openpyxl')
    writer.book = excelWorkbook

    whole = DF2.merge(DF3,on='Name',suffixes=("_(Eval 1)","_(Eval 2)"))
    
    cnt = 1
    lastpos = 3
    keyword_list = ['V Rdbk Accuracy','V Prog Accuracy','I Rdbk Accuracy','CurrentLimit Rdbk Accuracy','I_Lo Rdbk Accuracy','I Prog Accuracy',\
        'CurrentLimit Prog Accuracy','CurrentLoad','CurrentLine','VoltageLoad','VoltageLine','TransientResponse','CR Prog Accuracy','CP Prog Accuracy',\
        'CP Rdbk Accuracy','OvpAccuracyTest','NoiseTest','DOWN','UP','OVP Accuracy']
    DF2_temp = dfr.keyword_filter_row(DF2,'Name',keyword_list[0])
    DF2_temp = dfr.keyword_filter_row(DF2_temp,'Parent3','PowerSupply')
    DF3_temp = dfr.keyword_filter_row(DF3,'Name',keyword_list[0])
    DF3_temp = dfr.keyword_filter_row(DF3_temp,'Parent3','PowerSupply')
    result = DF2_temp.merge(DF3_temp,on='Name',suffixes=("_(Eval 1)","_(Eval 2)"))
    Name_list = result['Name'].tolist()
    result = result.rename(columns={'Name':'Voltage Readback Accuracy (PowerSupply)'})
    result.insert(DF2_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=cnt)
    worksheet = excelWorkbook['Data Analysis']
    col_names = ['Mean_(Eval 1)','Mean_(Eval 2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,cnt,result.shape[0],"Voltage Readback Accuracy",lastpos,result.shape[1] + 1) # width=17,height=2.6 # 27,25,28,59,57,60
    cnt += result.shape[0] + 2

    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[0],writer,'Voltage Readback Accuracy (ELoad)',True,'ELoad',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[1],writer,'Voltage Programming Accuracy (PowerSupply)',True,'PowerSupply',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[1],writer,'Voltage Programming Accuracy (ELoad)',True,'ELoad',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[2],writer,'Current Readback Accuracy (PowerSupply)',True,'PowerSupply',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[2],writer,'Current Readback Accuracy (ELoad)',True,'ELoad',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[5],writer,'Current Programming Accuracy (PowerSupply)',True,'PowerSupply',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[5],writer,'Current Programming Accuracy (ELoad)',True,'ELoad',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[4],writer,'I_Lo Rdbk Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keywords(DF2,DF3,[keyword_list[3],keyword_list[6]],writer,'CurrentLimit',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keywords(DF2,DF3,keyword_list[7:8+1],writer,'Current Load and Line Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keywords(DF2,DF3,keyword_list[9:10+1],writer,'Voltage Load and Line Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[11],writer,'Transient Response',False,startrowindex=cnt,lastpos=lastpos)
    DF2_temp = dfr.keywords_filter_row(DF2,'Name',keyword_list[12:13+1])
    DF2_temp = dfr.keyword_filter_row(DF2_temp,'Parent2',"ResistanceAccuracy")
    DF3_temp = dfr.keywords_filter_row(DF3,'Name',keyword_list[12:13+1])
    DF3_temp = dfr.keyword_filter_row(DF3_temp,'Parent2',"ResistanceAccuracy")
    DF2_temp['Name'] = DF2_temp['Name'].str.replace('?','Ω')
    DF3_temp['Name'] = DF3_temp['Name'].str.replace('?','Ω')
    result = DF2_temp.merge(DF3_temp,on='Name',suffixes=("_(Eval 1)","_(Eval 2)"))
    result = result.rename(columns={'Name':'CR & CP Programming Accuracy (Resistance)'})
    result.insert(DF2_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=cnt)
    col_names = ['Mean_(Eval 1)','Mean_(Eval 2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,cnt,result.shape[0],"CR & CP Programming Accuracy (Resistance)",lastpos,result.shape[1] + 1)
    cnt += result.shape[0] + 2
    DF2_temp = dfr.keywords_filter_row(DF2,'Name',keyword_list[12:13+1])
    DF2_temp = dfr.keyword_filter_row(DF2_temp,'Parent2',"PowerAccuracy")
    DF3_temp = dfr.keywords_filter_row(DF3,'Name',keyword_list[12:13+1])
    DF3_temp = dfr.keyword_filter_row(DF3_temp,'Parent2',"PowerAccuracy")
    result = DF2_temp.merge(DF3_temp,on='Name',suffixes=("_(Eval 1)","_(Eval 2)"))
    result = result.rename(columns={'Name':'CR & CP Programming Accuracy (Power Accuracy)'})
    result.insert(DF2_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=cnt)
    col_names = ['Mean_(Eval 1)','Mean_(Eval 2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,cnt,result.shape[0],"CR & CP Programming Accuracy (Power Accuracy)",lastpos,result.shape[1] + 1)
    cnt += result.shape[0] + 2
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[14],writer,'CP Readback Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[15],writer,'OvpAccuracyTest',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[16],writer,'NoiseTest',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[17],writer,'DOWN',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[18],writer,'UP',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(DF2,DF3,keyword_list[19],writer,'OVP Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    
    copy = whole
    whole = whole[~whole.Name.str.contains('|'.join(keyword_list))]
    if len(whole.index) != 0:
        whole = whole.rename(columns={'Name':'Others'})
        whole.insert(DF2_temp.shape[1], '','')
        whole.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=cnt)
        lastpos = plot_sigma_graph(worksheet,[27,59],cnt,whole.shape[0],"Others",lastpos)
        cnt += whole.shape[0] + 2

    writer.save()
    print("saving file ...")

    copy['Delta'] = abs(copy['Mean_(Eval 2)'] - copy['Mean_(Eval 1)'])
    iosf.dataframe_to_excel(copy,output_full_name,"Delta")
