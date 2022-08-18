import argparse, sys, os
import glob
import shutil
import pandas as pd
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.chart import (
    LineChart,
    Reference,
)
import data_df_repos as dfr
import io_save_file as iosf
import io_read_file as iorf
from datetime import datetime

def mover(source, destination):
    files = [x for x in os.listdir(source) if x.endswith((".csv",".xlsx"))]
    for file in range(len(files)):
        filename = files[file]
        try:
            shutil.move(source + "\\" + filename , destination + "\\" + filename)
        except FileNotFoundError:
            os.mkdir(destination)
            shutil.move(source + "\\" + filename , destination + "\\" + filename)
        except:
            pass

def plot_sigma_graph(worksheet,cols_list,start,length,title,rowpos,colpos,width=45,height=20):
    # Create a chart object.
    chart = LineChart()
    chart.width = width
    chart.height = height
    chart.title = title
    # Data
    start += 1
    for col in cols_list:
        data = Reference(worksheet, min_col=col, min_row=start, max_col=col, max_row=start + length + 1)
        chart.add_data(data, titles_from_data=True)
    name = Reference(worksheet, min_col=1, min_row=start, max_col=1, max_row=start + length + 1)
    chart.set_categories(name)
    chart.x_axis.tickLblPos = "low"
    # Insert the chart into the worksheet
    letter = get_column_letter(colpos)
    pos = letter + str(rowpos)
    rowpos = rowpos + 40
    worksheet.add_chart(chart,pos)
    return rowpos

def create_subtable_keyword(DF1,DF2,keyword_list,writer,title,Parent3,Parent3_word="",startrowindex=1,lastpos=1):
    DF1_temp = dfr.keyword_filter_row(DF1,'Name',keyword_list)
    DF2_temp = dfr.keyword_filter_row(DF2,'Name',keyword_list)
    if Parent3:
        DF1_temp = dfr.keyword_filter_row(DF1_temp,'Parent3_(Eval 1)',Parent3_word)
        DF2_temp = dfr.keyword_filter_row(DF2_temp,'Parent3_(Eval 1)',Parent3_word)
    result = DF1_temp.merge(DF2_temp,on='Name',suffixes=("_(1)","_(2)"))
    if len(result.index) == 0:
        return lastpos, startrowindex
    result = result.rename(columns={'Name':title})
    result.insert(DF1_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=startrowindex)
    col_names = ['Delta_(1)','Delta_(2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,startrowindex,result.shape[0],title,lastpos,result.shape[1] + 1)
    startrowindex += result.shape[0] + 2
    return lastpos, startrowindex

def create_subtable_keywords(DF1,DF2,keyword_list,writer,title,Parent3,Parent3_word="",startrowindex=1,lastpos=1):
    DF1_temp = dfr.keywords_filter_row(DF1,'Name',keyword_list)
    DF2_temp = dfr.keywords_filter_row(DF2,'Name',keyword_list)
    if Parent3:
        DF1_temp = dfr.keyword_filter_row(DF1_temp,'Parent3_(Eval 1)',Parent3_word)
        DF2_temp = dfr.keyword_filter_row(DF2_temp,'Parent3_(Eval 1)',Parent3_word)
    result = DF1_temp.merge(DF2_temp,on='Name',suffixes=("_(1)","_(2)"))
    if len(result.index) == 0:
        return lastpos, startrowindex
    result = result.rename(columns={'Name':title})
    result.insert(DF1_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=startrowindex)
    col_names = ['Delta_(1)','Delta_(2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,startrowindex,result.shape[0],title,lastpos,result.shape[1] + 1)
    startrowindex += result.shape[0] + 2
    return lastpos, startrowindex

input_file_loc = "C:\Local_Storage\Data"
source = input_file_loc + "\\Delta_input"
list_of_files = glob.glob(source + "\\*")
sorted_files = sorted(list_of_files, key=lambda t: -os.stat(t).st_mtime)
sorted_files = list(filter(lambda file: file.endswith(".xlsx"),sorted_files))

df1 = pd.read_excel(sorted_files[0],"Delta")
df1_name = name = os.path.basename(sorted_files[0])
df2 = pd.read_excel(sorted_files[1],"Delta")
df2_name = name = os.path.basename(sorted_files[1])

df1 = df1[['Name','Mean_(Eval 1)','Mean_(Eval 2)','Parent2_(Eval 1)','Parent3_(Eval 1)','Delta']]
df2 = df2[['Name','Mean_(Eval 1)','Mean_(Eval 2)','Parent2_(Eval 1)','Parent3_(Eval 1)','Delta']]

output_file_name = "Balsa_Delta_" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"
output_full_name = input_file_loc + '\\Delta_output' + '\\' + output_file_name

if not df1.empty and not df2.empty :
    excelWorkbook = Workbook()
    writer = pd.ExcelWriter(output_full_name, engine='openpyxl')
    writer.book = excelWorkbook

    whole = df1.merge(df2,on='Name',suffixes=("_(1)","_(2)"))
    
    cnt = 1
    lastpos = 3
    keyword_list = ['V Rdbk Accuracy','V Prog Accuracy','I Rdbk Accuracy','CurrentLimit Rdbk Accuracy','I_Lo Rdbk Accuracy','I Prog Accuracy',\
        'CurrentLimit Prog Accuracy','CurrentLoad','CurrentLine','VoltageLoad','VoltageLine','TransientResponse','CR Prog Accuracy','CP Prog Accuracy',\
        'CP Rdbk Accuracy','OvpAccuracyTest','NoiseTest','DOWN','UP','OVP Accuracy']
    df1_temp = dfr.keyword_filter_row(df1,'Name',keyword_list[0])
    df1_temp = dfr.keyword_filter_row(df1_temp,'Parent3_(Eval 1)','PowerSupply')
    df2_temp = dfr.keyword_filter_row(df2,'Name',keyword_list[0])
    df2_temp = dfr.keyword_filter_row(df2_temp,'Parent3_(Eval 1)','PowerSupply')
    result = df1_temp.merge(df2_temp,on='Name',suffixes=("_(1)","_(2)"))
    Name_list = result['Name'].tolist()
    result = result.rename(columns={'Name':'Voltage Readback Accuracy (PowerSupply)'})
    result.insert(df1_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=cnt)
    std=excelWorkbook.get_sheet_by_name('Sheet')
    excelWorkbook.remove_sheet(std)
    worksheet = excelWorkbook['Data Analysis']
    col_names = ['Delta_(1)','Delta_(2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,cnt,result.shape[0],"Voltage Readback Accuracy",lastpos,result.shape[1] + 1) # width=17,height=2.6 # 27,25,28,59,57,60
    cnt += result.shape[0] + 2

    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[0],writer,'Voltage Readback Accuracy (ELoad)',True,'ELoad',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[1],writer,'Voltage Programming Accuracy (PowerSupply)',True,'PowerSupply',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[1],writer,'Voltage Programming Accuracy (ELoad)',True,'ELoad',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[2],writer,'Current Readback Accuracy (PowerSupply)',True,'PowerSupply',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[2],writer,'Current Readback Accuracy (ELoad)',True,'ELoad',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[5],writer,'Current Programming Accuracy (PowerSupply)',True,'PowerSupply',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[5],writer,'Current Programming Accuracy (ELoad)',True,'ELoad',startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[4],writer,'I_Lo Rdbk Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keywords(df1,df2,[keyword_list[3],keyword_list[6]],writer,'CurrentLimit',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keywords(df1,df2,keyword_list[7:8+1],writer,'Current Load and Line Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keywords(df1,df2,keyword_list[9:10+1],writer,'Voltage Load and Line Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[11],writer,'Transient Response',False,startrowindex=cnt,lastpos=lastpos)
    df1_temp = dfr.keywords_filter_row(df1,'Name',keyword_list[12:13+1])
    df1_temp = dfr.keyword_filter_row(df1_temp,'Parent2_(Eval 1)',"ResistanceAccuracy")
    df2_temp = dfr.keywords_filter_row(df2,'Name',keyword_list[12:13+1])
    df2_temp = dfr.keyword_filter_row(df2_temp,'Parent2_(Eval 1)',"ResistanceAccuracy")
    df1_temp['Name'] = df1_temp['Name'].str.replace('?','Ω')
    df2_temp['Name'] = df2_temp['Name'].str.replace('?','Ω')
    result = df1_temp.merge(df2_temp,on='Name',suffixes=("_(1)","_(2)"))
    result = result.rename(columns={'Name':'CR & CP Programming Accuracy (Resistance)'})
    result.insert(df1_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=cnt)
    col_names = ['Delta_(1)','Delta_(2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,cnt,result.shape[0],"CR & CP Programming Accuracy (Resistance)",lastpos,result.shape[1] + 1)
    cnt += result.shape[0] + 2
    df1_temp = dfr.keywords_filter_row(df1,'Name',keyword_list[12:13+1])
    df1_temp = dfr.keyword_filter_row(df1_temp,'Parent2_(Eval 1)',"PowerAccuracy")
    df2_temp = dfr.keywords_filter_row(df2,'Name',keyword_list[12:13+1])
    df2_temp = dfr.keyword_filter_row(df2_temp,'Parent2_(Eval 1)',"PowerAccuracy")
    result = df1_temp.merge(df2_temp,on='Name',suffixes=("_(1)","_(2)"))
    result = result.rename(columns={'Name':'CR & CP Programming Accuracy (Power Accuracy)'})
    result.insert(df1_temp.shape[1], '','')
    result.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=cnt)
    col_names = ['Delta_(1)','Delta_(2)']
    col_list = []
    for name in col_names:
        col_list.append(list(result.columns).index(name) + 1)
    lastpos = plot_sigma_graph(worksheet,col_list,cnt,result.shape[0],"CR & CP Programming Accuracy (Power Accuracy)",lastpos,result.shape[1] + 1)
    cnt += result.shape[0] + 2
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[14],writer,'CP Readback Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[15],writer,'OvpAccuracyTest',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[16],writer,'NoiseTest',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[17],writer,'DOWN',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[18],writer,'UP',False,startrowindex=cnt,lastpos=lastpos)
    lastpos, cnt = create_subtable_keyword(df1,df2,keyword_list[19],writer,'OVP Accuracy',False,startrowindex=cnt,lastpos=lastpos)
    
    copy = whole
    whole = whole[~whole.Name.str.contains('|'.join(keyword_list))]
    if len(whole.index) != 0:
        whole = whole.rename(columns={'Name':'Others'})
        whole.insert(df1_temp.shape[1], '','')
        whole.to_excel(writer, sheet_name="Data Analysis", index=False, header=True, startrow=cnt)
        lastpos = plot_sigma_graph(worksheet,[27,59],cnt,whole.shape[0],"Others",lastpos)
        cnt += whole.shape[0] + 2

    writer.save()
    print("saving file ...")